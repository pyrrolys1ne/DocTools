"""PDF → Excel 表格提取测试。"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from doctools.errors import PDF_NO_TEXT, DoctoolsError
from doctools.pdf_excel import pdf_to_excel


def _make_table_pdf(path: Path) -> None:
    doc = SimpleDocTemplate(str(path))
    table = Table([["Name", "Qty", "Price"], ["Apple", "3", "12.5"], ["Pear", "5", "8.0"]])
    table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    doc.build([table])


def test_pdf_to_excel_extracts_table(tmp_path: Path) -> None:
    src = tmp_path / "in.pdf"
    dst = tmp_path / "out.xlsx"
    _make_table_pdf(src)

    note = pdf_to_excel(src, dst)

    assert dst.exists()
    wb = load_workbook(str(dst))
    assert len(wb.sheetnames) >= 1
    sheet = wb[wb.sheetnames[0]]
    # 首行表头应含 "Name" 或 "Apple"（reportlab 网格表格）
    cells = [str(c.value) for row in sheet.iter_rows() for c in row if c.value is not None]
    assert any("Apple" in c or "Name" in c for c in cells)
    assert note is None


def test_pdf_to_excel_scanned_pdf_reports_no_text(tmp_path: Path) -> None:
    """纯图片（无文字）PDF 报 PDF_NO_TEXT。"""
    from reportlab.pdfgen import canvas

    src = tmp_path / "scan.pdf"
    c = canvas.Canvas(str(src))
    c.rect(50, 50, 500, 700, fill=1, stroke=0)
    c.showPage()
    c.save()

    with pytest.raises(DoctoolsError) as excinfo:
        pdf_to_excel(src, tmp_path / "out.xlsx")

    assert excinfo.value.code == PDF_NO_TEXT
