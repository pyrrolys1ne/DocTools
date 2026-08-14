"""PDF → Excel 智能表格提取（PyMuPDF find_tables）。

提取电子 PDF 中的表格（有框/无框）到一个 xlsx，每张表一个独立 sheet
（命名 ``P{页}-T{表}``）。扫描件（全页无文字）报 ``PDF_NO_TEXT``。
"""

from __future__ import annotations

from pathlib import Path

from doctools.errors import PDF_NO_TEXT, DoctoolsError
from doctools.resource_policy import assert_pdf_pages


def _clean(cell: object) -> str:
    """单元格归一化：None 转空串，其余转 str。"""
    if cell is None:
        return ""
    return str(cell)


def pdf_to_excel(src: Path, dst: Path) -> str | None:
    """把单个 PDF 的表格提取到一个 xlsx（作为 process_batch 的 worker 使用）。

    返回附注（如"未检测到表格"），无附注时返回 None。
    """
    import fitz  # noqa: PLC0415  # PyMuPDF
    from openpyxl import Workbook  # noqa: PLC0415

    dst.parent.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(str(src))
    workbook = Workbook()
    workbook.remove(workbook.active)  # 删默认 sheet，只留实际提取的表
    sheet_count = 0
    try:
        assert_pdf_pages(len(doc))
        has_text = any(page.get_text().strip() for page in doc)
        if not has_text:
            raise DoctoolsError(
                PDF_NO_TEXT,
                "这个 PDF 没有可提取的文字，可能是扫描版图片 PDF，表格提取需要文字层。",
                "This PDF has no extractable text; table extraction requires a text layer.",
            )
        for page_index, page in enumerate(doc, start=1):
            tables = page.find_tables()
            for table_index, table in enumerate(tables.tables, start=1):
                rows = table.extract()
                if not rows:
                    continue
                sheet = workbook.create_sheet(title=f"P{page_index}-T{table_index}")
                for row in rows:
                    sheet.append([_clean(cell) for cell in row])
                sheet_count += 1
    finally:
        doc.close()

    if sheet_count == 0:
        sheet = workbook.create_sheet(title="说明")
        sheet.append(["未检测到表格。该 PDF 可能没有表格结构。"])
        workbook.save(str(dst))
        return "未检测到表格，已生成空说明。"
    workbook.save(str(dst))
    return None
